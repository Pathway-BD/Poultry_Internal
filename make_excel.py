from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 색상 팔레트 ──────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")
ROW_EVEN_FILL = PatternFill("solid", fgColor="EBF3FB")
ROW_ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
ACCENT_FILL   = PatternFill("solid", fgColor="D6E4F7")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
TITLE_FONT    = Font(name="Calibri", bold=True, size=13, color="1F3864")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="B0C4DE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill   = fill
    cell.font   = font
    cell.alignment = CENTER
    cell.border = BORDER

def style_cell(cell, even=True, align=LEFT):
    cell.fill      = ROW_EVEN_FILL if even else ROW_ODD_FILL
    cell.font      = BODY_FONT
    cell.alignment = align
    cell.border    = BORDER

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 : 제품 기본 정보
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "제품 기본 정보"

ws1.merge_cells("A1:G1")
ws1["A1"] = "Product Basic Information"
ws1["A1"].font      = TITLE_FONT
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 28

headers = ["제품명", "서브타이틀", "카테고리 태그", "권장 용량", "제품 설명", "주요 적용 종", "주요 적용 목적"]
for ci, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=ci, value=h)
    style_header(c)
ws1.row_dimensions[2].height = 22

products = [
    {
        "name":     "Lipidol",
        "subtitle": "The First Absorption Accelerator",
        "tags":     "cost, oil, fat, energy, ME, layer_prod, breeder_chick, dairy_milk, rumen",
        "dose":     "500g/MT feed",
        "desc":     "기능성 라이소포스포리피드(LPL) — 영양소 소화·흡수 가속화. M.V 100,000 kcal/kg. 비용 절감 및 생산성 향상제.",
        "species":  "Broiler / Layer / Breeder / Dairy / Aquaculture",
        "purpose":  "FCR 개선, 증체, 난생산성, 유생산성, 비용절감",
    },
    {
        "name":     "Endopower",
        "subtitle": "Optimal Cocktail NSP Enzyme for Feed Formulation",
        "tags":     "performance, fcr, enzyme, nutrient, digestibility, nsp, alt_protein",
        "dose":     "100-200g/MT feed",
        "desc":     "고체발효 NSPase 복합효소. M.V 500,000 kcal/kg. 옥수수-대두박 및 대체단백질(PKM, RSM) 사료에 최적.",
        "species":  "Broiler / Layer / Breeder / Aquaculture",
        "purpose":  "FCR 개선, 사료 소화율 향상, 에너지 절감",
    },
    {
        "name":     "ProBe-Bac",
        "subtitle": "New Preventive Approach for Bacterial Infection",
        "tags":     "gut, phage, enteritis, target_bacteria, wetlitter, agp_synergy, vertical_infection",
        "dose":     "500g-1,000g/MT feed 또는 WSP",
        "desc":     "살모넬라·대장균·클로스트리디움 등 58종 표적 박테리오파지. PAS 효과: 항생제와 병용 시 거의 완전한 균 제거.",
        "species":  "Broiler / Layer / Breeder / Aquaculture",
        "purpose":  "장건강, 세균성 질병 예방, 사망률 감소",
    },
    {
        "name":     "Q-Life",
        "subtitle": "Natural Alternative to Anticoccidials",
        "tags":     "immunity, coccidiosis, natural, anticoccidial",
        "dose":     "100-200g/MT feed",
        "desc":     "100% 천연 콕시듐 억제제. 여러 Eimeria 종 억제, 내성 없음.",
        "species":  "Broiler / Layer / Breeder",
        "purpose":  "콕시듐증 예방 및 면역 지원",
    },
    {
        "name":     "Growmax",
        "subtitle": "Essential Trace Mineral for Growing",
        "tags":     "performance, chromium, stress_resilience, mortality_reduce, all_species",
        "dose":     "500g/MT feed",
        "desc":     "유기 크로미움 — 스트레스 저항성 향상, 사망률 감소, 전 축종 성장 개선.",
        "species":  "Broiler / Layer / Breeder / Dairy / Aquaculture",
        "purpose":  "성장·FCR 개선, 스트레스 저항, 사망률 감소",
    },
    {
        "name":     "YeaMune-UP",
        "subtitle": "Immune Booster",
        "tags":     "immunity, MOS, beta-glucan, gut, stress, rumen, dairy_milk, yeast_cell_wall",
        "dose":     "300-500g/MT feed",
        "desc":     "S. cerevisiae boulardii 기반 β-글루칸 & MOS. 면역 강화, 장건강, 낙농 반추위 개선.",
        "species":  "Broiler / Layer / Breeder / Dairy",
        "purpose":  "면역 강화, 백신 반응 개선, 열 스트레스 대응",
    },
    {
        "name":     "Genikan",
        "subtitle": "High-quality Yeast Culture — Rich in Amino Acids, Vitamins & UGF",
        "tags":     "gut, palatability, yeast, feed intake, fermentation, early_gut, all_species",
        "dose":     "1-2kg/MT feed",
        "desc":     "옥수수·대두박 발효 효모 배양물. 초기 사육 및 사료 전환기에 최적. 전 축종 장건강 빠른 회복.",
        "species":  "Broiler / Layer / Breeder / Dairy / Aquaculture",
        "purpose":  "기호성 향상, 장건강, 초기 성장 지원",
    },
    {
        "name":     "FermKito",
        "subtitle": "Optimized for Egg Quality & Liver Function",
        "tags":     "gut, fatty_liver, cholesterol, chitosan, layer_egg, shell_quality, aqua_liver",
        "dose":     "1-3kg/MT feed",
        "desc":     "발효 키토산 & COS — 면역조절. 난각 강화, 파란율 감소, 지방간 증후군 억제 (채란계·수산).",
        "species":  "Layer / Aquaculture",
        "purpose":  "난각 품질, 지방간 억제, 간 기능 개선",
    },
    {
        "name":     "StreX",
        "subtitle": "Anti-Stress Solution",
        "tags":     "cooling, heat, feed_intake, anti_stress, immunity",
        "dose":     "500g/MT",
        "desc":     "열 스트레스·백신 스트레스·질병 도전에 대한 종합 항스트레스 솔루션. 냉각 감각 및 사료 섭취량 지원.",
        "species":  "Broiler / Layer / Breeder",
        "purpose":  "열 스트레스 완화, 백신 스트레스, 사망률 감소",
    },
]

for ri, p in enumerate(products, 3):
    even = ri % 2 == 0
    vals = [p["name"], p["subtitle"], p["tags"], p["dose"], p["desc"], p["species"], p["purpose"]]
    for ci, v in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=v)
        style_cell(c, even=even)
        if ci == 1:
            c.font = Font(name="Calibri", bold=True, size=10)
    ws1.row_dimensions[ri].height = 45

widths1 = [14, 30, 40, 18, 55, 32, 32]
for i, w in enumerate(widths1, 1):
    set_col_width(ws1, i, w)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 : 제품별 추천 적용 상황 (reasonMap 기반)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("추천 적용 상황")

ws2.merge_cells("A1:F1")
ws2["A1"] = "Product Recommendation Map"
ws2["A1"].font      = TITLE_FONT
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 28

headers2 = ["제품명", "Detail 키", "한글 상황명", "적용 Species", "추천 이유 (영문)", "우선순위"]
for ci, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    style_header(c)
ws2.row_dimensions[2].height = 22

detail_kr = {
    "heat":          "열 스트레스",
    "vaccination":   "백신 스트레스",
    "disease":       "질병 도전",
    "mortality":     "폐사율",
    "fcr":           "FCR 개선",
    "weight":        "증체",
    "uniformity":    "균일도",
    "egg_production":"산란율 & 난중",
    "egg_quality":   "난각 품질",
    "liver_function":"간 기능",
    "chick_quality": "병아리 품질",
    "milk_yield":    "유생산량",
    "rumen_health":  "반추위 건강",
    "growth":        "성장률 (수산)",
    "general_perf":  "일반 생산성",
    "general_gut":   "일반 장건강",
    "general_cost":  "일반 비용절감",
}

detail_species = {
    "heat":          "Broiler / Layer / Breeder / Dairy",
    "vaccination":   "Broiler / Layer / Breeder",
    "disease":       "Broiler / Layer / Breeder / Aquaculture",
    "mortality":     "Broiler / Layer / Breeder / Dairy",
    "fcr":           "Broiler / Layer",
    "weight":        "Broiler",
    "uniformity":    "Broiler",
    "egg_production":"Layer / Breeder",
    "egg_quality":   "Layer",
    "liver_function":"Layer",
    "chick_quality": "Breeder",
    "milk_yield":    "Dairy",
    "rumen_health":  "Dairy",
    "growth":        "Aquaculture",
    "general_perf":  "All",
    "general_gut":   "Dairy / Aquaculture",
    "general_cost":  "Dairy / Aquaculture",
}

reason_map = [
    # heat
    ("heat",          "StreX",      "Cooling sensation technology directly combats heat stress, maintaining feed intake during high temperatures.", 1),
    ("heat",          "YeaMune-UP", "β-glucan & MOS strengthen immune defense weakened by heat stress.", 2),
    ("heat",          "Growmax",    "Organic chromium improves stress resilience and reduces heat-related mortality.", 3),
    # vaccination
    ("vaccination",   "StreX",      "Supports birds through vaccination stress, maintaining feed intake and reducing post-vaccination performance drop.", 1),
    ("vaccination",   "YeaMune-UP", "Immune modulation via β-glucan helps birds respond better to vaccination.", 2),
    # disease
    ("disease",       "ProBe-Bac",  "Bacteriophage directly eliminates pathogenic bacteria (Salmonella, E.coli, C.perfringens) — 58 target strains without resistance.", 1),
    ("disease",       "StreX",      "Broad anti-stress support strengthens overall disease resistance.", 2),
    ("disease",       "Q-Life",     "Natural anticoccidial — if coccidiosis is the primary disease challenge.", 3),
    ("disease",       "YeaMune-UP", "MOS & β-glucan enhance innate immunity against disease challenges.", 4),
    # mortality
    ("mortality",     "ProBe-Bac",  "Targeted bacteriophage eliminates pathogenic bacteria that are the leading cause of mortality. PAS effect with AGP for maximum control.", 1),
    ("mortality",     "StreX",      "Anti-stress formula reduces stress-related mortality across all conditions.", 2),
    ("mortality",     "Growmax",    "Organic chromium proven to reduce mortality rate and improve stress resilience.", 3),
    # fcr
    ("fcr",           "Lipidol",    "LPL technology accelerates fat & nutrient absorption — directly improves FCR. M.V 100,000 kcal/kg.", 1),
    ("fcr",           "Endopower",  "NSPase enzyme unlocks trapped energy from corn-SBM, improving feed efficiency. M.V 500,000 kcal/kg.", 2),
    ("fcr",           "Growmax",    "Organic chromium enhances glucose metabolism, supporting lean growth and better FCR.", 3),
    # weight
    ("weight",        "Lipidol",    "Accelerated nutrient digestion & absorption drives faster weight gain.", 1),
    ("weight",        "Growmax",    "Chromium improves nutrient partitioning — more energy directed to muscle growth.", 2),
    ("weight",        "Endopower",  "Multi-enzyme complex maximizes energy extraction from feed ingredients.", 3),
    # uniformity
    ("uniformity",    "Lipidol",    "Consistent nutrient absorption across the flock improves uniformity.", 1),
    ("uniformity",    "Endopower",  "Better feed digestion reduces variability in nutrient uptake between birds.", 2),
    ("uniformity",    "Genikan",    "Great palatability ensures consistent feed intake across the flock.", 3),
    # egg_production
    ("egg_production","Lipidol",    "Enhanced fat absorption increases energy available for egg production & size.", 1),
    ("egg_production","Endopower",  "Maximizes ME from feed — more energy directed to egg production.", 2),
    ("egg_production","FermKito",   "Liver health support sustains peak laying performance.", 3),
    # egg_quality
    ("egg_quality",   "FermKito",   "Chitosan & COS strengthen egg shell, reduce breakage rate.", 1),
    ("egg_quality",   "Lipidol",    "Better nutrient absorption delivers more calcium & minerals to shell formation.", 2),
    ("egg_quality",   "Genikan",    "Gut health improvement ensures optimal mineral absorption for shell quality.", 3),
    # liver_function
    ("liver_function","FermKito",   "Directly targets fatty liver syndrome — reduces cholesterol deposition.", 1),
    ("liver_function","Lipidol",    "LPLs improve fat metabolism, reducing liver fat accumulation.", 2),
    ("liver_function","Genikan",    "Yeast metabolites support liver detoxification and recovery.", 3),
    # chick_quality
    ("chick_quality", "Lipidol",    "Improved maternal nutrition through LPLs enhances egg quality and chick vitality.", 1),
    ("chick_quality", "Endopower",  "Better nutrient extraction in breeder feed translates to healthier chicks.", 2),
    ("chick_quality", "Growmax",    "Chromium in breeder diet improves reproductive performance and hatchability.", 3),
    # milk_yield
    ("milk_yield",    "Lipidol",    "LPLs enhance fat absorption in rumen bypass, increasing energy for milk production.", 1),
    ("milk_yield",    "YeaMune-UP", "Yeast cell wall & MOS improve rumen microbiome, boosting milk yield.", 2),
    ("milk_yield",    "Genikan",    "Yeast culture enhances rumen fermentation and feed palatability.", 3),
    # rumen_health
    ("rumen_health",  "YeaMune-UP", "β-glucan & MOS optimize rumen microbiome for better fermentation efficiency.", 1),
    ("rumen_health",  "Lipidol",    "Improved lipid utilization supports rumen health and energy balance.", 2),
    ("rumen_health",  "Genikan",    "Yeast metabolites stimulate beneficial rumen bacteria.", 3),
    # growth (aqua)
    ("growth",        "Lipidol",    "Enhanced nutrient absorption drives faster growth in aquaculture species.", 1),
    ("growth",        "Endopower",  "Multi-enzyme complex improves feed utilization and growth rate.", 2),
    ("growth",        "Growmax",    "Organic chromium supports lean growth and stress resilience.", 3),
    # general_perf
    ("general_perf",  "Lipidol",    "LPL technology — the foundational performance enhancer. M.V 100,000 kcal/kg.", 1),
    ("general_perf",  "Growmax",    "Organic chromium for growth, stress resilience, and mortality reduction.", 2),
    ("general_perf",  "Endopower",  "NSPase enzyme maximizes energy from corn-SBM diets. M.V 500,000 kcal/kg.", 3),
    # general_gut
    ("general_gut",   "Lipidol",    "LPL technology supports gut integrity and nutrient absorption efficiency.", 1),
    ("general_gut",   "Genikan",    "Yeast metabolites promote beneficial gut microbiome balance and digestive health.", 2),
    ("general_gut",   "ProBe-Bac",  "Bacteriophage eliminates pathogenic bacteria, restoring gut microbial balance.", 3),
    # general_cost
    ("general_cost",  "Lipidol",    "LPL technology improves nutrient utilization, enabling feed cost reduction without performance loss.", 1),
    ("general_cost",  "Genikan",    "Yeast culture enhances feed palatability and digestibility, improving feed efficiency.", 2),
    ("general_cost",  "ProBe-Bac",  "Bacteriophage replaces costly antibiotics, reducing treatment costs while maintaining gut health.", 3),
]

# 제품명 기준으로 그룹 색상 구분
prod_order = ["Lipidol","Endopower","ProBe-Bac","Q-Life","Growmax","YeaMune-UP","Genikan","FermKito","StreX"]
prod_fills = [
    PatternFill("solid", fgColor="DDEEFF"),
    PatternFill("solid", fgColor="E8F5E9"),
    PatternFill("solid", fgColor="FFF3E0"),
    PatternFill("solid", fgColor="F3E5F5"),
    PatternFill("solid", fgColor="FCE4EC"),
    PatternFill("solid", fgColor="E0F7FA"),
    PatternFill("solid", fgColor="FFFDE7"),
    PatternFill("solid", fgColor="F1F8E9"),
    PatternFill("solid", fgColor="FBE9E7"),
]
prod_fill_map = {p: f for p, f in zip(prod_order, prod_fills)}

# 제품명 기준으로 정렬
reason_map_sorted = sorted(reason_map, key=lambda x: (prod_order.index(x[1]) if x[1] in prod_order else 99, x[3]))

ri = 3
for detail, prod, reason, rank in reason_map_sorted:
    fill = prod_fill_map.get(prod, ROW_ODD_FILL)
    vals = [prod, detail, detail_kr.get(detail, detail), detail_species.get(detail, "-"), reason, rank]
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.fill      = fill
        c.font      = BODY_FONT
        c.alignment = LEFT if ci not in (6,) else CENTER
        c.border    = BORDER
        if ci == 1:
            c.font = Font(name="Calibri", bold=True, size=10)
    ws2.row_dimensions[ri].height = 40
    ri += 1

widths2 = [14, 16, 18, 30, 65, 8]
for i, w in enumerate(widths2, 1):
    set_col_width(ws2, i, w)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 : 실험·분석 데이터 개요
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("실험·분석 데이터 개요")

ws3.merge_cells("A1:E1")
ws3["A1"] = "Experimental & Analysis Data Overview"
ws3["A1"].font      = TITLE_FONT
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 28

headers3 = ["분류", "주요 관련 제품", "데이터 내용", "입력 파라미터", "산출 결과"]
for ci, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=ci, value=h)
    style_header(c)
ws3.row_dimensions[2].height = 22

analysis_data = [
    (
        "Lipidol 사료비 최적화\n(Feed Cost Calculator)",
        "Lipidol",
        "3가지 시나리오별 사료 배합 최적화 시뮬레이션\n"
        "A. Full Fat Soya 감량\n"
        "B. Soybean Oil 대체\n"
        "C. FFS + SYO 동시 감량",
        "Full Fat Soya (kg, 조지방%, 조단백%, AME)\n"
        "Soybean Oil (L, AME)\n"
        "SBM 조단백%, 사료 총량(kg)\n"
        "Lipidol AME 값",
        "시나리오별 Lipidol 포함량(g)\n"
        "원료 감량량(kg 또는 L)\n"
        "조지방·조단백 손실량\n"
        "SBM 보충 필요량\n"
        "ME 추가분(kcal/kg)\n"
        "비용 절감액",
    ),
    (
        "ProBe-Bac 급수 처리 스케줄\n(Water Treatment Schedule)",
        "ProBe-Bac",
        "일령별 급수 탱크 투여 스케줄 자동 계산\n"
        "보충 주기 및 1회 투여량 산출",
        "품종(Ross308 / Cobb500)\n"
        "일령(일)\n"
        "사육 수수\n"
        "음수:사료 비율(mg/L)\n"
        "PPM 용량\n"
        "탱크 용량(L)",
        "일별 사료 섭취량\n"
        "일별 음수량(L)\n"
        "보충 필요일 & 투여량(g)\n"
        "누적 처리 소요량",
    ),
    (
        "첨가제 FCR 개선 시뮬레이션\n(Additive Impact Simulation)",
        "전 제품 공통",
        "목표 FCR 설정 시 사료비 절감 효과 및\n비용 대비 효익(ROI) 계산",
        "품종(Ross308 / Cobb500)\n"
        "일령(일)\n"
        "목표 FCR\n"
        "사육 수수\n"
        "첨가제 용량(g/MT)\n"
        "단가 및 통화",
        "표준 FCR vs 목표 FCR 비교\n"
        "절감 사료량(kg)\n"
        "첨가제 총 비용\n"
        "순 비용 절감액\n"
        "ROI",
    ),
    (
        "육계 성장 성적 분석\n(Broiler Performance Analysis)",
        "공통 (성적 벤치마킹)",
        "품종별(Ross308, Cobb500) 표준 성장 곡선 대비\n실제 농장 성적 비교 분석\n\n"
        "▸ 표준 데이터 범위: 0~40일령, 41개 구간\n"
        "▸ Ross308: 0일 44g → 40일 2,798g (FCR 1.493)\n"
        "▸ Cobb500: 0일 42g → 40일 3,062g (FCR 1.522)",
        "품종\n일령\n실측 체중(g)\n실측 FCR\n초생추 중량\n사료 섭취량(일별/누적)",
        "체중 편차(실측 vs 표준)\n"
        "FCR 편차\n"
        "일별 증체량 곡선\n"
        "사료 섭취량 곡선\n"
        "신호등(🟢🟡🔴) 진단",
    ),
    (
        "채란계 성적 분석\n(Layer Performance Analysis)",
        "공통 (성적 벤치마킹)",
        "6개 품종 표준 성적 대비 실농장 추적 분석\n\n"
        "▸ Hisex Brown       (육성 1~18주, 산란 18~100주)\n"
        "▸ Hyline Brown 2025 (육성 1~17주, 산란 18~100주)\n"
        "▸ Hyline W80        (육성 1~17주, 산란 18~72주)\n"
        "▸ Hisex White       (산란기 전체)\n"
        "▸ Lohmann Brown Classic (육성+산란)\n"
        "▸ Lohmann LSL Lite  (육성+산란)",
        "품종\n주령(육성/산란기)\n실측 체중\n사료 섭취량\n산란율(%)\n난중(g)\n생존율(%)",
        "표준 vs 실측 성적 비교표\n"
        "산란율·난중·사료 효율 편차\n"
        "누적 산란수·누적 난괴량\n"
        "100주까지 성적 곡선\n"
        "신호등 진단",
    ),
]

fill_list = [
    PatternFill("solid", fgColor="DDEEFF"),
    PatternFill("solid", fgColor="FFF3E0"),
    PatternFill("solid", fgColor="E8F5E9"),
    PatternFill("solid", fgColor="F3E5F5"),
    PatternFill("solid", fgColor="E0F7FA"),
]

for ri, (cat, prod, content, inputs, outputs) in enumerate(analysis_data, 3):
    fill = fill_list[(ri - 3) % len(fill_list)]
    for ci, v in enumerate([cat, prod, content, inputs, outputs], 1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.fill      = fill
        c.font      = BODY_FONT
        c.alignment = LEFT
        c.border    = BORDER
    ws3.row_dimensions[ri].height = 100

widths3 = [28, 22, 45, 40, 40]
for i, w in enumerate(widths3, 1):
    set_col_width(ws3, i, w)


# ════════════════════════════════════════════════════════════════════════════
# 저장
# ════════════════════════════════════════════════════════════════════════════
out = "/home/user/Poultry_Internal/Product_Selector_정리.xlsx"
wb.save(out)
print("saved:", out)
